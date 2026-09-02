# JOB 6b — the Record writing sheet: one tab per day, Records 002-006.
# Reads _prefill.json (exported from reveal/record-entries.mjs). Writes nothing
# back to the repo.
import json, io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

PRE = {e["no"]: e for e in json.load(io.open(r"C:\AI\_night-20260811\_prefill.json", encoding="utf-8"))}

DAYS = [
    (2, "2026-08-18", "TUESDAY",  "Day 2"),
    (3, "2026-08-19", "WEDNESDAY","Day 3"),
    (4, "2026-08-20", "THURSDAY", "Day 4"),
    (5, "2026-08-21", "FRIDAY",   "Day 5"),
    (6, "2026-08-22", "SATURDAY", "Day 6"),
]

INK   = "FF211F1C"; DIM = "FF57544D"
HEAD  = PatternFill("solid", fgColor="FFE2DED3")
LABEL = PatternFill("solid", fgColor="FFEFECE3")
FIELD = PatternFill("solid", fgColor="FFFFFFFF")
PRESET= PatternFill("solid", fgColor="FFFFF9E0")   # already written
thin  = Side(style="thin", color="FFC6C2B7")
BOX   = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook(); wb.remove(wb.active)

def sect_label(e, i):
    try: return e["sections"][i].get("label") or ""
    except Exception: return ""

def sect_body(e, i):
    try:
        s = e["sections"][i]
        return "\n\n".join(s.get("body") or s.get("paras") or [])
    except Exception: return ""

for no, date, dow, dayname in DAYS:
    e = PRE.get(no, {})
    ws = wb.create_sheet(f"{dayname} - Record {no:03d}")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 96
    ws.column_dimensions["C"].width = 22

    row = [1]
    def put(label, value, note, fill=FIELD, height=None, counter=None, wrap=True):
        r = row[0]
        a = ws.cell(r, 1, label); a.font = Font(bold=True, size=10, color=INK)
        a.alignment = Alignment(vertical="top", wrap_text=True); a.fill = LABEL; a.border = BOX
        b = ws.cell(r, 2, value); b.font = Font(size=11, color=INK)
        b.alignment = Alignment(vertical="top", wrap_text=wrap); b.border = BOX
        b.fill = PRESET if value else fill
        c = ws.cell(r, 3, note); c.font = Font(size=9, color=DIM, italic=True)
        c.alignment = Alignment(vertical="top", wrap_text=True); c.border = BOX
        if counter:
            ws.cell(r, 4, counter).font = Font(size=9, bold=True, color=DIM)
        if height: ws.row_dimensions[r].height = height
        row[0] = r + 1

    # ---- masthead -----------------------------------------------------------
    t = ws.cell(row[0], 1, f"RECORD {no:03d}   ·   {dow} {date}   ·   {dayname}")
    t.font = Font(bold=True, size=15, color=INK); t.fill = HEAD; t.border = BOX
    ws.cell(row[0], 2, "").fill = HEAD; ws.cell(row[0], 2).border = BOX
    ws.cell(row[0], 3, "").fill = HEAD; ws.cell(row[0], 3).border = BOX
    ws.row_dimensions[row[0]].height = 26; row[0] += 1
    g = ws.cell(row[0], 1, "Type in the WHITE cells. Cream cells are already written — change them or leave them.")
    g.font = Font(size=10, italic=True, color=DIM); row[0] += 2

    # ---- the fields ---------------------------------------------------------
    put("THE HEADLINE", e.get("title") or "",
        "One line. 62 characters maximum — the count is in column D and it is live. Longer than that and the index row overflows.",
        height=30, counter=f"=IF(B{row[0]}=\"\",\"\",LEN(B{row[0]})&\" / 62\"&IF(LEN(B{row[0]})>62,\"  TOO LONG\",\"\"))")

    deck = e.get("line") or ""
    put("THE DECK — line 1", deck.split("\n")[0] if deck else "",
        "First of two lines under the headline. Keep it short enough not to wrap.",
        height=30)
    put("THE DECK — line 2", (deck.split("\n")[1] if "\n" in deck else ""),
        "Second line. BOTH LINES TOGETHER must be 130 characters or fewer — column D counts both.",
        height=30, counter=f"=IF(AND(B{row[0]-1}=\"\",B{row[0]}=\"\"),\"\",LEN(B{row[0]-1})+LEN(B{row[0]})+1&\" / 130\"&IF(LEN(B{row[0]-1})+LEN(B{row[0]})+1>130,\"  TOO LONG\",\"\"))")

    ws.cell(row[0], 1, "").border = None; row[0] += 1
    h = ws.cell(row[0], 1, "THE SECTIONS"); h.font = Font(bold=True, size=12, color=INK); h.fill = HEAD; h.border = BOX
    ws.cell(row[0], 2, "A label, then its paragraphs. Leave a BLANK LINE between paragraphs — that is what splits them.").font = Font(size=9, italic=True, color=DIM)
    ws.cell(row[0], 2).fill = HEAD; ws.cell(row[0], 2).border = BOX; ws.cell(row[0], 2).alignment = Alignment(vertical="top", wrap_text=True)
    ws.cell(row[0], 3, "").fill = HEAD; ws.cell(row[0], 3).border = BOX
    ws.row_dimensions[row[0]].height = 24; row[0] += 1

    existing = len(e.get("sections") or [])
    for i in range(max(6, existing)):
        lab = sect_label(e, i); bod = sect_body(e, i)
        put(f"Section {i+1} — LABEL", lab,
            "Short. CAPITALS if you want it to read as a heading. No limit." if i == 0 else "No limit. Leave blank to stop here.",
            height=22)
        put(f"Section {i+1} — TEXT", bod,
            "No limit at all. Blank line between paragraphs. Anything in { curly braces } is a note to Ops and never reaches a visitor.",
            height=132 if bod else 96)
        ws.cell(row[0], 1, ""); row[0] += 1

    put("ATTACHMENTS", "", "One per line. A filename, a link, or a plain description. Leave blank if there are none.", height=60)
    ws.freeze_panes = "A4"

# ---- the front sheet --------------------------------------------------------
front = wb.create_sheet("READ ME FIRST", 0)
front.sheet_view.showGridLines = False
front.column_dimensions["A"].width = 118
LINES = [
    ("THE RECORD — Days 2 to 6", 16, True),
    ("", 11, False),
    ("One tab per day, along the bottom. Type into the white cells. Cream cells already have", 11, False),
    ("something in them — that is text already in the museum. Change it or leave it.", 11, False),
    ("", 11, False),
    ("THE TWO LIMITS, AND THEY ARE COUNTED FOR YOU IN COLUMN D", 12, True),
    ("    THE HEADLINE   62 characters.  Longer, and the index row overflows.", 11, False),
    ("    THE DECK       130 characters across BOTH lines.  It is the whole summary; nothing is cut.", 11, False),
    ("Column D turns to TOO LONG the moment you go over. Nothing else on the sheet has a limit.", 11, False),
    ("", 11, False),
    ("HOW A SECTION WORKS", 12, True),
    ("    A LABEL, then its TEXT.  Leave a BLANK LINE between paragraphs — that is what splits them.", 11, False),
    ("    Six section slots per day. Use as many as you need and leave the rest empty.", 11, False),
    ("    To type a new line inside a cell, press ALT+ENTER.", 11, False),
    ("", 11, False),
    ("NOTES TO OPS GO IN CURLY BRACES", 12, True),
    ("    Anything inside { } is a note to me and never reaches a visitor. A gate fails the build", 11, False),
    ("    if a brace ever survives to launch, so you can write in them freely.", 11, False),
    ("", 11, False),
    ("WHAT IS ALREADY HERE", 12, True),
    ("    Record 002 and 003 have a headline, a deck and their sections.", 11, False),
    ("    Record 004 and 005 have an EXECUTIVE SUMMARY and nothing else — no headline, no deck.", 11, False),
    ("    Record 006 is empty. It does not exist in the museum yet.", 11, False),
    ("", 11, False),
    ("WHEN YOU ARE DONE", 12, True),
    ("    Save it and tell me. I read this file and land it in the tree. You do not touch the code.", 11, False),
]
rr = 2
for text, size, bold in LINES:
    c = front.cell(rr, 1, text)
    c.font = Font(size=size, bold=bold, color=INK if bold else DIM)
    c.alignment = Alignment(vertical="top", wrap_text=False)
    rr += 1

OUT = r"C:\AI\_night-20260811\RECORD_days-2-to-6.xlsx"
wb.save(OUT)
print("wrote", OUT, "with", len(wb.sheetnames), "sheets:", wb.sheetnames)
